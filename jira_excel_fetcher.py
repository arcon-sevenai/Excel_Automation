# import os
# import io
# import json
# import pandas as pd
# import requests
# from openpyxl import load_workbook
# from requests.auth import HTTPBasicAuth

# # ============================================================
# # LOAD CONFIG (JIRA Credentials)
# # ============================================================
# with open("config.json", "r") as f:
#     config = json.load(f)

# JIRA_URL = config["jira"]["base_url"]
# EMAIL = config["jira"]["email"]
# API_TOKEN = config["jira"]["api_token"]
# AUTH = HTTPBasicAuth(EMAIL, API_TOKEN)
# HEADERS = {"Accept": "application/json"}

# # ============================================================
# # VALIDATE EXCEL STRUCTURE
# # ============================================================
# def validate_excel_structure(file_bytes: bytes) -> bool:
#     """
#     Validate Excel file structure.
#     Ensures that:
#       - It opens properly
#       - At least one sheet has headers
#       - No embedded images (indicating wrong format)
#     """
#     try:
#         xls = pd.ExcelFile(io.BytesIO(file_bytes))
#         wb = load_workbook(io.BytesIO(file_bytes))

#         # Check for embedded images
#         for sheet in wb.sheetnames:
#             ws = wb[sheet]
#             if getattr(ws, "_images", []):
#                 print(f" Sheet '{sheet}' has images — invalid format.")
#                 return False

#         # Check for valid headers
#         for sheet in xls.sheet_names:
#             df = pd.read_excel(xls, sheet_name=sheet)
#             if df.empty:
#                 continue
#             headers = [str(h).strip().lower() for h in df.columns]
#             if any(h for h in headers if h and not h.startswith("unnamed")):
#                 print(f" Sheet '{sheet}' has valid headers.")
#                 return True

#         print(" No valid headers found in any sheet.")
#         return False

#     except Exception as e:
#         print(f" Validation failed: {e}")
#         return False

# # ============================================================
# # FETCH EXCEL ATTACHMENTS FROM JIRA
# # ============================================================
# def fetch_excel_from_jira(ticket_id: str):
#     """
#     Fetch Excel attachments for a given JIRA ticket.
#     Saves valid files to a local folder named after the ticket.
#     Returns a list of file paths.
#     """
#     print(f"\n Fetching attachments for JIRA ticket: {ticket_id}")
#     issue_url = f"{JIRA_URL}/rest/api/3/issue/{ticket_id}?fields=attachment"
#     resp = requests.get(issue_url, headers=HEADERS, auth=AUTH)

#     if resp.status_code == 404:
#         print(" Ticket not found or invalid permissions.")
#         return []
#     resp.raise_for_status()

#     attachments = resp.json()["fields"].get("attachment", [])
#     excel_files = [a for a in attachments if a["filename"].lower().endswith((".xls", ".xlsx", ".csv"))]

#     if not excel_files:
#         print(" No Excel attachments found for this ticket.")
#         return []

#     ticket_folder = os.path.join("uploads", ticket_id)
#     os.makedirs(ticket_folder, exist_ok=True)

#     valid_files = []
#     for att in excel_files:
#         filename = att["filename"]
#         print(f" Downloading: {filename}")
#         file_resp = requests.get(att["content"], headers=HEADERS, auth=AUTH)
#         file_resp.raise_for_status()
#         file_bytes = file_resp.content

#         if not validate_excel_structure(file_bytes):
#             print(f" '{filename}' failed validation — skipped.")
#             continue

#         file_path = os.path.join(ticket_folder, filename)
#         with open(file_path, "wb") as f:
#             f.write(file_bytes)
#         print(f" Saved valid Excel → {file_path}")
#         valid_files.append(file_path)

#     if not valid_files:
#         print(" No valid Excel files after validation.")
#     else:
#         print(f" All valid files saved in: {ticket_folder}")

#     return valid_files


# if __name__ == "__main__":
#     ticket_id = input(" Enter JIRA Ticket ID (e.g., PAM-1234): ").strip()
#     if not ticket_id:
#         print(" No ticket ID provided. Exiting.")
#         exit(0)
#     fetch_excel_from_jira(ticket_id)











# # jira_excel_fetcher.py
# import os
# import io
# import re
# import json
# import unicodedata
# import pandas as pd
# import requests
# from openpyxl import load_workbook
# from requests.auth import HTTPBasicAuth

# # ============================================================
# # LOAD CONFIG (JIRA Credentials)
# # ============================================================
# with open("config.json", "r", encoding="utf-8") as f:
#     config = json.load(f)

# JIRA_URL = config["jira"]["base_url"]
# EMAIL = config["jira"]["email"]
# API_TOKEN = config["jira"]["api_token"]
# AUTH = HTTPBasicAuth(EMAIL, API_TOKEN)
# HEADERS = {"Accept": "application/json"}

# # ============================================================
# # CONSTANTS / HELPERS
# # ============================================================
# ALLOWED_EXTS = (".xls", ".xlsx", ".xlsm", ".csv")

# def _safe_name(s: str, fallback: str = "sheet") -> str:
#     """Return a filesystem-safe string."""
#     if not s:
#         s = fallback
#     s = unicodedata.normalize("NFKD", s)
#     s = re.sub(r'[\\/:*?"<>|]', "_", s)           # illegal FS chars
#     s = re.sub(r"\s+", " ", s).strip()            # collapse whitespace
#     return s or fallback

# def _detect_header_and_data(df_no_header: pd.DataFrame):
#     """
#     Heuristic header detection:
#       - Header row = first row with ≥2 non-empty cells.
#       - Must have at least one non-empty row below header.
#     Returns: (header_row_idx or None, has_data_after_header: bool)
#     """
#     header_row_idx = None
#     for i, row in df_no_header.iterrows():
#         if row.count() >= 2:
#             header_row_idx = i
#             break
#     if header_row_idx is None:
#         return None, False
#     below = df_no_header.loc[header_row_idx + 1 :]
#     has_data = below.notna().any(axis=1).any()
#     return header_row_idx, has_data

# def _sheet_has_only_images_openpyxl(wb, sheet_name: str) -> bool:
#     """True if the sheet has embedded images and effectively no data values."""
#     try:
#         ws = wb[sheet_name]
#         has_images = bool(getattr(ws, "_images", []))
#         # Any cell value present?
#         has_values = False
#         for row in ws.iter_rows(values_only=True):
#             if any(cell is not None and str(cell).strip() != "" for cell in row):
#                 has_values = True
#                 break
#         return has_images and not has_values
#     except Exception:
#         # If inspection fails, don't claim images-only
#         return False

# def _export_df_to_single_sheet_xlsx(out_xlsx_path: str, df_no_header: pd.DataFrame):
#     """Export DataFrame (no header) to a single-sheet XLSX file."""
#     out_dir = os.path.dirname(out_xlsx_path)
#     os.makedirs(out_dir, exist_ok=True)
#     with pd.ExcelWriter(out_xlsx_path, engine="openpyxl") as writer:
#         df_no_header.to_excel(writer, index=False, header=False, sheet_name="Sheet1")

# def _make_output_file_path(ticket_id: str, attachment_folder: str,
#                            attachment_name: str, sheet_idx_one_based: int,
#                            sheet_name: str, decision_prefix: str) -> str:
#     """
#     Build the single-file output path INSIDE the attachment folder:
#       uploads/<ticket_id>/<attachment_base>/
#         <ticket_id>__<attachment_base>__<decision_prefix>_sheetK__<sheetname>.xlsx

#     decision_prefix: "right" or "wrong"
#     """
#     attachment_base = _safe_name(os.path.splitext(attachment_name)[0], fallback="attachment")
#     sheet_safe = _safe_name(sheet_name, fallback=f"sheet{sheet_idx_one_based}")
#     base_tag = f"{ticket_id}__{attachment_base}__{decision_prefix}_sheet{sheet_idx_one_based}__{sheet_safe}"
#     out_xlsx = os.path.join(attachment_folder, f"{base_tag}.xlsx")
#     return out_xlsx, base_tag

# # ============================================================
# # PER-FILE, PER-SHEET INTAKE (classify sheets & export)
# # ============================================================
# def intake_workbook(file_bytes: bytes, ticket_id: str, ticket_folder: str, attachment_name: str) -> list:
#     """
#     Inspect each sheet in the attachment independently.
#       - If sheet is plausible (header+data somewhere) -> export a RIGHT file (xlsx).
#       - If sheet is images-only / empty / no header+data / load error -> export a WRONG file (xlsx or error note).
#     Exports each sheet as a single file (no per-sheet subfolder) into:
#       uploads/<ticket_id>/<attachment_base>/<ticket>__<attachment_base>__right_sheetK__<sheetname>.xlsx

#     Returns: list of RIGHT-sheet XLSX paths to pass to module 2.
#     Writes an _intake_manifest.json under the attachment folder.
#     """
#     right_paths = []
#     manifest = {
#         "ticket_id": ticket_id,
#         "attachment": attachment_name,
#         "sheets": []
#     }

#     ext = os.path.splitext(attachment_name)[1].lower()

#     # Attachment folder
#     attachment_base = _safe_name(os.path.splitext(attachment_name)[0], fallback="attachment")
#     attachment_folder = os.path.join(ticket_folder, attachment_base)
#     os.makedirs(attachment_folder, exist_ok=True)

#     # For images-only check (xlsx/xlsm only)
#     wb = None
#     if ext in (".xlsx", ".xlsm"):
#         try:
#             wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
#         except Exception:
#             wb = None  # non-fatal

#     # CSV handling: treat as single "sheet"
#     if ext == ".csv":
#         try:
#             try:
#                 df = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding="utf-8")
#             except UnicodeDecodeError:
#                 df = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding="latin-1")

#             header_row_idx, has_data = _detect_header_and_data(df)
#             if header_row_idx is not None and has_data:
#                 decision = "right"
#                 reason = "header_found"
#             else:
#                 decision = "wrong"
#                 reason = "no_header_or_data"

#             out_xlsx, base_tag = _make_output_file_path(
#                 ticket_id, attachment_folder, attachment_name, 1, "CSV", decision
#             )
#             _export_df_to_single_sheet_xlsx(out_xlsx, df)

#             if decision == "right":
#                 right_paths.append(out_xlsx)

#             manifest["sheets"].append({
#                 "sheet_index": 1,
#                 "sheet_name": "CSV",
#                 "decision": decision.upper(),
#                 "reason": reason,
#                 "images_only": False,
#                 "header_row_guess": int(header_row_idx) if header_row_idx is not None else None,
#                 "export_path": out_xlsx,
#                 "export_tag": base_tag,
#                 "attachment_folder": attachment_folder
#             })

#         except Exception as e:
#             # record failure (write an error note file in the attachment folder)
#             base_tag = f"{ticket_id}__{attachment_base}__wrong_sheet1__CSV"
#             try:
#                 with open(os.path.join(attachment_folder, f"{base_tag}__error.txt"), "w", encoding="utf-8") as ef:
#                     ef.write(f"csv_load_error: {e}")
#             except Exception:
#                 pass

#             manifest["sheets"].append({
#                 "sheet_index": 1,
#                 "sheet_name": "CSV",
#                 "decision": "SKIP",
#                 "reason": f"csv_load_error: {e}",
#                 "images_only": False,
#                 "header_row_guess": None,
#                 "export_path": None,
#                 "export_tag": base_tag,
#                 "attachment_folder": attachment_folder
#             })

#         # write manifest at the attachment level
#         with open(os.path.join(attachment_folder, "_intake_manifest.json"), "w", encoding="utf-8") as mf:
#             json.dump(manifest, mf, indent=2, ensure_ascii=False)
#         return right_paths

#     # Excel (xls/xlsx/xlsm)
#     try:
#         xls = pd.ExcelFile(io.BytesIO(file_bytes))
#         sheet_names = xls.sheet_names
#     except Exception as e:
#         # Can't open workbook with pandas: record at attachment level and bail
#         with open(os.path.join(attachment_folder, "_intake_manifest.json"), "w", encoding="utf-8") as mf:
#             json.dump({"error": f"workbook_open_error: {e}"}, mf, indent=2, ensure_ascii=False)
#         return right_paths

#     for idx, sheet in enumerate(sheet_names, start=1):
#         try:
#             df_raw = pd.read_excel(xls, sheet_name=sheet, header=None)

#             images_only = False
#             if wb is not None and sheet in wb.sheetnames:
#                 images_only = _sheet_has_only_images_openpyxl(wb, sheet)

#             header_row_idx, has_data = _detect_header_and_data(df_raw)

#             if images_only:
#                 decision = "wrong"
#                 reason = "images_only"
#             elif header_row_idx is not None and has_data:
#                 decision = "right"
#                 reason = "header_found"
#             else:
#                 decision = "wrong"
#                 reason = "no_header_or_data"

#             out_xlsx, base_tag = _make_output_file_path(
#                 ticket_id, attachment_folder, attachment_name, idx, str(sheet), decision
#             )
#             _export_df_to_single_sheet_xlsx(out_xlsx, df_raw)

#             if decision == "right":
#                 right_paths.append(out_xlsx)

#             manifest["sheets"].append({
#                 "sheet_index": idx,
#                 "sheet_name": str(sheet),
#                 "decision": decision.upper(),
#                 "reason": reason,
#                 "images_only": bool(images_only),
#                 "header_row_guess": int(header_row_idx) if header_row_idx is not None else None,
#                 "export_path": out_xlsx,
#                 "export_tag": base_tag,
#                 "attachment_folder": attachment_folder
#             })

#         except Exception as e:
#             # Per-sheet failure -> write an error note in the attachment folder
#             base_tag = f"{ticket_id}__{attachment_base}__wrong_sheet{idx}__{_safe_name(str(sheet))}"
#             try:
#                 with open(os.path.join(attachment_folder, f"{base_tag}__error.txt"), "w", encoding="utf-8") as ef:
#                     ef.write(f"sheet_load_error: {e}")
#             except Exception:
#                 pass

#             manifest["sheets"].append({
#                 "sheet_index": idx,
#                 "sheet_name": str(sheet),
#                 "decision": "SKIP",
#                 "reason": f"sheet_load_error: {e}",
#                 "images_only": False,
#                 "header_row_guess": None,
#                 "export_path": None,
#                 "export_tag": base_tag,
#                 "attachment_folder": attachment_folder
#             })

#     # Write manifest for this attachment
#     with open(os.path.join(attachment_folder, "_intake_manifest.json"), "w", encoding="utf-8") as mf:
#         json.dump(manifest, mf, indent=2, ensure_ascii=False)

#     return right_paths

# # ============================================================
# # FETCH EXCEL ATTACHMENTS FROM JIRA
# # ============================================================
# def fetch_excel_from_jira(ticket_id: str):
#     """
#     Fetch Excel/CSV attachments for a given JIRA ticket and split each sheet into its own file,
#     laid out as:

#       uploads/<ticket_id>/<attachment_base>/
#         <ticket_id>__<attachment_base>__right_sheet1__<sheetname>.xlsx
#         <ticket_id>__<attachment_base>__wrong_sheet2__<sheetname>.xlsx
#         _intake_manifest.json

#     Returns a list of file paths (RIGHT sheets only) to pass to module 2.
#     """
#     print(f"\n Fetching attachments for JIRA ticket: {ticket_id}")
#     issue_url = f"{JIRA_URL}/rest/api/3/issue/{ticket_id}?fields=attachment"
#     resp = requests.get(issue_url, headers=HEADERS, auth=AUTH)

#     if resp.status_code == 404:
#         print(" Ticket not found or invalid permissions.")
#         return []
#     resp.raise_for_status()

#     attachments = resp.json()["fields"].get("attachment", [])
#     files = [a for a in attachments if a["filename"].lower().endswith(ALLOWED_EXTS)]

#     if not files:
#         print(" No Excel/CSV attachments found for this ticket.")
#         return []

#     # Parent folder: uploads/<ticket_id>/
#     ticket_folder = os.path.join("uploads", ticket_id)
#     os.makedirs(ticket_folder, exist_ok=True)

#     right_sheet_paths = []

#     for att in files:
#         filename = att["filename"]
#         print(f" Downloading: {filename}")
#         file_resp = requests.get(att["content"], headers=HEADERS, auth=AUTH)
#         file_resp.raise_for_status()
#         file_bytes = file_resp.content

#         # Save original attachment for audit
#         attachment_base = _safe_name(os.path.splitext(filename)[0], fallback="attachment")
#         attachment_folder = os.path.join(ticket_folder, attachment_base)
#         os.makedirs(attachment_folder, exist_ok=True)
#         original_path = os.path.join(attachment_folder, filename)
#         with open(original_path, "wb") as f:
#             f.write(file_bytes)
#         print(f" Saved original → {original_path}")

#         # Per-sheet intake & export (files placed directly in attachment folder)
#         per_sheet_right = intake_workbook(file_bytes, ticket_id, ticket_folder, filename)
#         print(f"  -> Included sheets for {filename}: {len(per_sheet_right)}")
#         right_sheet_paths.extend(per_sheet_right)

#     if not right_sheet_paths:
#         print(" No valid sheets detected after per-sheet inspection.")
#     else:
#         print(f" All per-sheet exports completed under: {ticket_folder}")
#         print(f" Right-sheet files to pass to module 2: {len(right_sheet_paths)}")

#     return right_sheet_paths

# # ============================================================
# # MAIN
# # ============================================================
# # if __name__ == "__main__":
# #     ticket_id = input(" Enter JIRA Ticket ID (e.g., PAM-1234): ").strip()
# #     if not ticket_id:
# #         print(" No ticket ID provided. Exiting.")
# #         raise SystemExit(0)
# #     fetch_excel_from_jira(ticket_id)



























# jira_excel_fetcher.py
import os
import io
import re
import json
import unicodedata
import pandas as pd
import requests
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth

# ============================================================
# LOAD CONFIG (JIRA Credentials)
# ============================================================
with open("config.json", "r", encoding="utf-8") as f:
    config = json.load(f)

JIRA_URL = config["jira"]["base_url"]
EMAIL = config["jira"]["email"]
API_TOKEN = config["jira"]["api_token"]
AUTH = HTTPBasicAuth(EMAIL, API_TOKEN)
HEADERS = {"Accept": "application/json"}

# ============================================================
# CONSTANTS / HELPERS
# ============================================================
ALLOWED_EXTS = (".xls", ".xlsx", ".xlsm", ".csv")

def _safe_name(s: str, fallback: str = "sheet") -> str:
    """Return a filesystem-safe string."""
    if not s:
        s = fallback
    s = unicodedata.normalize("NFKD", s)
    s = re.sub(r'[\\/:*?"<>|]', "_", s)           # illegal FS chars
    s = re.sub(r"\s+", " ", s).strip()            # collapse whitespace
    return s or fallback

def _detect_header_and_data(df_no_header: pd.DataFrame):
    """
    Heuristic header detection:
      - Header row = first row with ≥2 non-empty cells.
      - Must have at least one non-empty row below header.
    Returns: (header_row_idx or None, has_data_after_header: bool)
    """
    header_row_idx = None
    for i, row in df_no_header.iterrows():
        if row.count() >= 2:
            header_row_idx = i
            break
    if header_row_idx is None:
        return None, False
    below = df_no_header.loc[header_row_idx + 1 :]
    has_data = below.notna().any(axis=1).any()
    return header_row_idx, has_data

def _sheet_has_only_images_openpyxl(wb, sheet_name: str) -> bool:
    """True if the sheet has embedded images and effectively no data values."""
    try:
        ws = wb[sheet_name]
        has_images = bool(getattr(ws, "_images", []))
        # Any cell value present?
        has_values = False
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                has_values = True
                break
        return has_images and not has_values
    except Exception:
        # If inspection fails, don't claim images-only
        return False

def _export_df_to_single_sheet_xlsx(out_xlsx_path: str, df_no_header: pd.DataFrame):
    """Export DataFrame (no header) to a single-sheet XLSX file."""
    out_dir = os.path.dirname(out_xlsx_path)
    os.makedirs(out_dir, exist_ok=True)
    with pd.ExcelWriter(out_xlsx_path, engine="openpyxl") as writer:
        df_no_header.to_excel(writer, index=False, header=False, sheet_name="Sheet1")

def _make_output_file_path(ticket_id: str, attachment_folder: str,
                           attachment_name: str, sheet_idx_one_based: int,
                           sheet_name: str, decision_prefix: str) -> str:
    """
    Build the single-file output path INSIDE the attachment folder:
      uploads/<ticket_id>/<attachment_base>/
        <ticket_id>__<attachment_base>__<decision_prefix>_sheetK__<sheetname>.xlsx

    decision_prefix: "right" or "wrong"
    """
    attachment_base = _safe_name(os.path.splitext(attachment_name)[0], fallback="attachment")
    sheet_safe = _safe_name(sheet_name, fallback=f"sheet{sheet_idx_one_based}")
    base_tag = f"{ticket_id}__{attachment_base}__{decision_prefix}_sheet{sheet_idx_one_based}__{sheet_safe}"
    out_xlsx = os.path.join(attachment_folder, f"{base_tag}.xlsx")
    return out_xlsx, base_tag

# Small helper to produce friendly reason text for header/data checks
def _friendly_header_reason(header_row_idx, has_data):
    if header_row_idx is None:
        return "No usable header row found (need at least two non-empty cells in a header row)."
    if not has_data:
        return "Header found but no data rows below—skipping sheet."
    return "Header found."

# ============================================================
# PER-FILE, PER-SHEET INTAKE (classify sheets & export)
# ============================================================
def intake_workbook(file_bytes: bytes, ticket_id: str, ticket_folder: str, attachment_name: str) -> list:
    """
    Inspect each sheet in the attachment independently.
      - If sheet is plausible (header+data somewhere) -> export a RIGHT file (xlsx).
      - If sheet is images-only / empty / no header+data / load error -> export a WRONG file (xlsx or error note).
    Exports each sheet as a single file (no per-sheet subfolder) into:
      uploads/<ticket_id>/<attachment_base>/<ticket>__<attachment_base>__right_sheetK__<sheetname>.xlsx

    Returns: list of RIGHT-sheet XLSX paths to pass to module 2.
    Writes an _intake_manifest.json under the attachment folder.
    """
    right_paths = []
    manifest = {
        "ticket_id": ticket_id,
        "attachment": attachment_name,
        "sheets": []
    }

    ext = os.path.splitext(attachment_name)[1].lower()

    # Attachment folder
    attachment_base = _safe_name(os.path.splitext(attachment_name)[0], fallback="attachment")
    attachment_folder = os.path.join(ticket_folder, attachment_base)
    os.makedirs(attachment_folder, exist_ok=True)

    # For images-only check (xlsx/xlsm only)
    wb = None
    if ext in (".xlsx", ".xlsm"):
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            wb = None  # non-fatal
            print(f'⚠️ Could not open workbook with openpyxl for image inspection: {e}')

    # CSV handling: treat as single "sheet"
    if ext == ".csv":
        try:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding="utf-8")
            except UnicodeDecodeError:
                df = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding="latin-1")

            header_row_idx, has_data = _detect_header_and_data(df)
            if header_row_idx is not None and has_data:
                decision = "right"
                reason = "header_found"
                reason_detail = _friendly_header_reason(header_row_idx, has_data)
                print(f"✅ CSV header OK: {reason_detail}")
            else:
                decision = "wrong"
                reason = "no_header_or_data"
                reason_detail = _friendly_header_reason(header_row_idx, has_data)
                print(f"🚫 CSV not usable: {reason_detail}")

            out_xlsx, base_tag = _make_output_file_path(
                ticket_id, attachment_folder, attachment_name, 1, "CSV", decision
            )
            _export_df_to_single_sheet_xlsx(out_xlsx, df)

            if decision == "right":
                right_paths.append(out_xlsx)

            manifest["sheets"].append({
                "sheet_index": 1,
                "sheet_name": "CSV",
                "decision": decision.upper(),
                "reason": reason,
                "reason_detail": reason_detail,
                "images_only": False,
                "header_row_guess": int(header_row_idx) if header_row_idx is not None else None,
                "export_path": out_xlsx,
                "export_tag": base_tag,
                "attachment_folder": attachment_folder
            })

        except Exception as e:
            # record failure (write an error note file in the attachment folder)
            base_tag = f"{ticket_id}__{attachment_base}__wrong_sheet1__CSV"
            msg = f"CSV couldn't be parsed (encoding/format issue). Underlying error: {e}"
            try:
                with open(os.path.join(attachment_folder, f"{base_tag}__error.txt"), "w", encoding="utf-8") as ef:
                    ef.write(f"csv_load_error: {e}")
            except Exception:
                pass
            print(f" {msg}")

            manifest["sheets"].append({
                "sheet_index": 1,
                "sheet_name": "CSV",
                "decision": "SKIP",
                "reason": "csv_load_error",
                "reason_detail": "CSV couldn't be parsed (encoding/format issue).",
                "images_only": False,
                "header_row_guess": None,
                "export_path": None,
                "export_tag": base_tag,
                "attachment_folder": attachment_folder
            })

        # write manifest at the attachment level
        with open(os.path.join(attachment_folder, "_intake_manifest.json"), "w", encoding="utf-8") as mf:
            json.dump(manifest, mf, indent=2, ensure_ascii=False)
        return right_paths

    # Excel (xls/xlsx/xlsm)
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        sheet_names = xls.sheet_names
    except Exception as e:
        # Can't open workbook with pandas: record at attachment level and bail
        print(" Workbook couldn't be opened (corrupted, password-protected, or unsupported).")
        with open(os.path.join(attachment_folder, "_intake_manifest.json"), "w", encoding="utf-8") as mf:
            json.dump({"error": f"workbook_open_error: {e}"}, mf, indent=2, ensure_ascii=False)
        return right_paths

    for idx, sheet in enumerate(sheet_names, start=1):
        try:
            df_raw = pd.read_excel(xls, sheet_name=sheet, header=None)

            images_only = False
            if wb is not None and sheet in wb.sheetnames:
                images_only = _sheet_has_only_images_openpyxl(wb, sheet)

            header_row_idx, has_data = _detect_header_and_data(df_raw)

            if images_only:
                decision = "wrong"
                reason = "images_only"
                reason_detail = "Sheet has only embedded images and no cell data—skipping."
                print(f" [{sheet}] {reason_detail}")
            elif header_row_idx is not None and has_data:
                decision = "right"
                reason = "header_found"
                reason_detail = _friendly_header_reason(header_row_idx, has_data)
                print(f" [{sheet}] {reason_detail}")
            else:
                decision = "wrong"
                reason = "no_header_or_data"
                reason_detail = _friendly_header_reason(header_row_idx, has_data)
                print(f" [{sheet}] {reason_detail}")

            out_xlsx, base_tag = _make_output_file_path(
                ticket_id, attachment_folder, attachment_name, idx, str(sheet), decision
            )
            _export_df_to_single_sheet_xlsx(out_xlsx, df_raw)

            if decision == "right":
                right_paths.append(out_xlsx)

            manifest["sheets"].append({
                "sheet_index": idx,
                "sheet_name": str(sheet),
                "decision": decision.upper(),
                "reason": reason,
                "reason_detail": reason_detail,
                "images_only": bool(images_only),
                "header_row_guess": int(header_row_idx) if header_row_idx is not None else None,
                "export_path": out_xlsx,
                "export_tag": base_tag,
                "attachment_folder": attachment_folder
            })

        except Exception as e:
            # Per-sheet failure -> write an error note in the attachment folder
            base_tag = f"{ticket_id}__{attachment_base}__wrong_sheet{idx}__{_safe_name(str(sheet))}"
            try:
                with open(os.path.join(attachment_folder, f"{base_tag}__error.txt"), "w", encoding="utf-8") as ef:
                    ef.write(f"sheet_load_error: {e}")
            except Exception:
                pass
            print(f" [{sheet}] Sheet couldn't be parsed (merged cells/format issue). Underlying error: {e}")

            manifest["sheets"].append({
                "sheet_index": idx,
                "sheet_name": str(sheet),
                "decision": "SKIP",
                "reason": "sheet_load_error",
                "reason_detail": "Sheet couldn't be parsed (merged cells/format issue).",
                "images_only": False,
                "header_row_guess": None,
                "export_path": None,
                "export_tag": base_tag,
                "attachment_folder": attachment_folder
            })

    # Write manifest for this attachment
    with open(os.path.join(attachment_folder, "_intake_manifest.json"), "w", encoding="utf-8") as mf:
        json.dump(manifest, mf, indent=2, ensure_ascii=False)

    return right_paths

# ============================================================
# FETCH EXCEL ATTACHMENTS FROM JIRA
# ============================================================
def fetch_excel_from_jira(ticket_id: str):
    """
    Fetch Excel/CSV attachments for a given JIRA ticket and split each sheet into its own file,
    laid out as:

      uploads/<ticket_id>/<attachment_base>/
        <ticket_id>__<attachment_base>__right_sheet1__<sheetname>.xlsx
        <ticket_id>__<attachment_base>__wrong_sheet2__<sheetname>.xlsx
        _intake_manifest.json

    Returns a list of file paths (RIGHT sheets only) to pass to module 2.
    """
    print(f"\n Fetching attachments for JIRA ticket: {ticket_id}")
    issue_url = f"{JIRA_URL}/rest/api/3/issue/{ticket_id}?fields=attachment"

    try:
        resp = requests.get(issue_url, headers=HEADERS, auth=AUTH)
    except requests.RequestException as e:
        print(f" Network error talking to JIRA: {e}")
        return []

    if resp.status_code == 404:
        print(" Ticket not found or you don't have permission to view this issue.")
        return []

    try:
        resp.raise_for_status()
    except requests.HTTPError as e:
        print(f" JIRA request failed (HTTP {resp.status_code}): {resp.reason}")
        return []

    data = resp.json()
    attachments = (data.get("fields") or {}).get("attachment", [])
    if attachments is None:
        attachments = []

    files = [a for a in attachments if a.get("filename", "").lower().endswith(ALLOWED_EXTS)]

    if not attachments:
        print("ℹ️ This issue has no attachments.")
    if not files:
        print(" No Excel/CSV attachments found for this ticket.")
        return []

    # Parent folder: uploads/<ticket_id>/
    ticket_folder = os.path.join("uploads", ticket_id)
    os.makedirs(ticket_folder, exist_ok=True)

    right_sheet_paths = []

    for att in files:
        filename = att.get("filename", "attachment")
        content_url = att.get("content")

        if not content_url:
            print(f" Attachment '{filename}' has no downloadable content URL—skipping.")
            continue

        print(f" Downloading: {filename}")
        try:
            file_resp = requests.get(content_url, headers=HEADERS, auth=AUTH)
        except requests.RequestException as e:
            print(f" Failed to download attachment '{filename}': {e}")
            continue

        if file_resp.status_code != 200:
            print(f" Failed to download attachment '{filename}': HTTP {file_resp.status_code}")
            continue

        file_bytes = file_resp.content

        # Save original attachment for audit
        try:
            attachment_base = _safe_name(os.path.splitext(filename)[0], fallback="attachment")
            attachment_folder = os.path.join(ticket_folder, attachment_base)
            os.makedirs(attachment_folder, exist_ok=True)
            original_path = os.path.join(attachment_folder, filename)
            with open(original_path, "wb") as f:
                f.write(file_bytes)
            print(f" Saved original → {original_path}")
        except OSError as e:
            print(f" Couldn't save original attachment '{filename}': {e}")
            # Even if we can't save, we can still try to process in-memory:
            attachment_folder = os.path.join(ticket_folder, _safe_name(os.path.splitext(filename)[0], fallback="attachment"))
            os.makedirs(attachment_folder, exist_ok=True)

        # Per-sheet intake & export (files placed directly in attachment folder)
        per_sheet_right = intake_workbook(file_bytes, ticket_id, ticket_folder, filename)
        print(f"  -> Included sheets for {filename}: {len(per_sheet_right)}")
        right_sheet_paths.extend(per_sheet_right)

    if not right_sheet_paths:
        print(" No valid sheets detected after per-sheet inspection.")
    else:
        print(f" All per-sheet exports completed under: {ticket_folder}")
        print(f" Right-sheet files to pass to module 2: {len(right_sheet_paths)}")

    return right_sheet_paths

# ============================================================
# MAIN
# ============================================================
# if __name__ == "__main__":
#     ticket_id = input(" Enter JIRA Ticket ID (e.g., PAM-1234): ").strip()
#     if not ticket_id:
#         print(" No ticket ID provided. Exiting.")
#         raise SystemExit(0)
#     fetch_excel_from_jira(ticket_id)
